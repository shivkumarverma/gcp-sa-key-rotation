"""
Build an .xlsx report of service account key scan/rotation results.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.rotator import (
    AUTO_ROTATE_THRESHOLD_DAYS,
    EXPIRING_SOON_THRESHOLD_DAYS,
    OK_THRESHOLD_DAYS,
    RotationRecord,
)

# ---------------------------------------------------------------------------
# Palette — light pastel row tints with bold, deeper status text
# ---------------------------------------------------------------------------
_STATUS_STYLES = {
    "OK":            {"row": "ECFDF5", "badge": "D1FAE5", "font": "047857"},  # mint / emerald
    "Rotated":       {"row": "ECFDF5", "badge": "D1FAE5", "font": "047857"},
    "Expiring Soon": {"row": "FEF9C3", "badge": "FEF08A", "font": "A16207"},  # pale lemon / amber
    "Critical":      {"row": "FFEDD5", "badge": "FED7AA", "font": "C2410C"},  # pale peach / orange
    "Very Critical": {"row": "FEE2E2", "badge": "FECACA", "font": "B91C1C"},  # pale rose / red
    "Error":         {"row": "FEE2E2", "badge": "FECACA", "font": "991B1B"},
}

# Header (darker than the body, as requested)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="175CD3")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Segoe UI Semibold", size=11)

# Body text
_BODY_FONT      = Font(name="Segoe UI", size=10, color="1F2937")
_BODY_FONT_MUTE = Font(name="Segoe UI", size=10, color="6B7280")

# Title banner (auto-rotate notice)
_NOTICE_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")
_NOTICE_FONT = Font(bold=True, color="92400E", name="Segoe UI Semibold", size=11)

# Legend strip
_LEGEND_FILL = PatternFill(fill_type="solid", fgColor="F1F5F9")
_LEGEND_FONT = Font(italic=True, color="475467", name="Segoe UI", size=10)

# Zebra stripe for rows with no status tint (defensive fallback)
_ZEBRA_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")

# Borders
_BODY_BORDER = Border(bottom=Side(border_style="thin", color="E5E7EB"))
_HEADER_BORDER = Border(bottom=Side(border_style="medium", color="1E3A8A"))

# Validation cell (subtle but readable)
_VALIDATION_PASS_FILL = PatternFill(fill_type="solid", fgColor="D1FAE5")
_VALIDATION_FAIL_FILL = PatternFill(fill_type="solid", fgColor="FECACA")
_VALIDATION_PASS_FONT = Font(bold=True, name="Segoe UI", size=10, color="047857")
_VALIDATION_FAIL_FONT = Font(bold=True, name="Segoe UI", size=10, color="991B1B")

_COLUMNS = [
    ("Project Name",        28),
    ("Project ID",          28),
    ("Service Account",     42),
    ("Expiry Date",         22),
    ("Days Remaining",      16),
    ("Status",              16),
    ("Storage Location",    54),
    ("Rotation Timestamp",  22),
    ("Error",               40),
    ("Key Validation",      16),
]

_KEY_VALIDATION_COL = len(_COLUMNS)  # 1-based index of the Key Validation column
_STATUS_COL = 6
_DAYS_COL = 5


def _paint_row(ws, row_idx: int, last_col: int, fill: PatternFill) -> None:
    """Apply a uniform background fill across every cell in a row."""
    for col_idx in range(1, last_col + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def build_report(
    records: list[RotationRecord],
    output_path: Path,
    rotation_enabled: bool,
) -> Path:
    """Write an .xlsx report and return its path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SA Key Rotation Report" if rotation_enabled else "SA Key Scan Report"
    ws.sheet_view.showGridLines = False

    last_col = len(_COLUMNS)

    # ----- Row 1: auto-rotate notice banner ---------------------------------
    notice_cell = ws.cell(
        row=1, column=1,
        value=(
            f"  ⚠  Notice — keys within {AUTO_ROTATE_THRESHOLD_DAYS} days of expiry will "
            f"auto rotate on the next scheduled run (when ENABLE_ROTATION is on)."
        ),
    )
    notice_cell.font = _NOTICE_FONT
    notice_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.row_dimensions[1].height = 28
    _paint_row(ws, 1, last_col, _NOTICE_FILL)

    # ----- Row 2: status legend ---------------------------------------------
    legend_text = (
        f"Status bands  •  OK > {OK_THRESHOLD_DAYS}d   "
        f"•  Expiring Soon {EXPIRING_SOON_THRESHOLD_DAYS + 1}–{OK_THRESHOLD_DAYS}d   "
        f"•  Critical {AUTO_ROTATE_THRESHOLD_DAYS + 1}–{EXPIRING_SOON_THRESHOLD_DAYS}d   "
        f"•  Very Critical ≤ {AUTO_ROTATE_THRESHOLD_DAYS}d  (auto-rotate)"
    )
    legend_cell = ws.cell(row=2, column=1, value=legend_text)
    legend_cell.font = _LEGEND_FONT
    legend_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.row_dimensions[2].height = 22
    _paint_row(ws, 2, last_col, _LEGEND_FILL)

    # ----- Row 3: spacer (visual breathing room) ----------------------------
    ws.row_dimensions[3].height = 6

    # ----- Row 4: column headers --------------------------------------------
    header_row = 4
    for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = _HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = f"A{header_row + 1}"

    # ----- Data rows --------------------------------------------------------
    align_left = Alignment(horizontal="left", vertical="center", indent=1)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left_wrap = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=False)

    for offset, rec in enumerate(records):
        row_idx = header_row + 1 + offset

        expiry_str = rec.expiry_date.strftime("%Y-%m-%d %H:%M UTC") if rec.expiry_date else "N/A"
        days_str = str(rec.days_remaining) if rec.days_remaining is not None else "N/A"
        rot_ts_str = rec.rotation_timestamp.strftime("%Y-%m-%d %H:%M UTC") if rec.rotation_timestamp else ""

        if rec.key_valid is None:
            validation_str = ""
        else:
            validation_str = "Pass" if rec.key_valid else "Fail"

        row_values = [
            rec.project_name or rec.project_id,
            rec.project_id,
            rec.sa_email,
            expiry_str,
            days_str,
            rec.status,
            rec.storage_location,
            rot_ts_str,
            rec.error_message,
            validation_str,
        ]

        style = _STATUS_STYLES.get(rec.status)
        row_fill = PatternFill(fill_type="solid", fgColor=style["row"]) if style else _ZEBRA_FILL

        col_alignments = {
            1: align_left,
            2: align_left,
            3: align_left,
            4: align_center,
            5: align_center,
            6: align_center,
            7: align_left_wrap,
            8: align_center,
            9: align_left_wrap,
            10: align_center,
        }

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.alignment = col_alignments.get(col_idx, align_left)
            cell.border = _BODY_BORDER
            cell.fill = row_fill

            # Status badge — bold coloured text in a slightly darker tint
            if col_idx == _STATUS_COL and style:
                cell.font = Font(bold=True, name="Segoe UI Semibold", size=10, color=style["font"])
                cell.fill = PatternFill(fill_type="solid", fgColor=style["badge"])

            # Days Remaining — bold + coloured to echo the status band
            elif col_idx == _DAYS_COL and style and rec.days_remaining is not None:
                cell.font = Font(bold=True, name="Segoe UI", size=10, color=style["font"])

            # Storage location / error — muted text when empty / informational
            elif col_idx in (7, 9) and not value:
                cell.font = _BODY_FONT_MUTE

            # Validation column override
            if col_idx == _KEY_VALIDATION_COL and rec.key_valid is not None:
                if rec.key_valid:
                    cell.fill = _VALIDATION_PASS_FILL
                    cell.font = _VALIDATION_PASS_FONT
                else:
                    cell.fill = _VALIDATION_FAIL_FILL
                    cell.font = _VALIDATION_FAIL_FONT

        ws.row_dimensions[row_idx].height = 22

    # ----- Auto-filter on the data table ------------------------------------
    last_data_row = max(header_row, header_row + len(records))
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_data_row}"

    wb.save(output_path)
    return output_path
